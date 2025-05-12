# -*- coding: utf-8 -*-
# @Author  : 
# @Desc    : 奖励函数分析工具

from typing import Dict, List, Any, Tuple
import numpy as np
import pandas as pd
from openpyxl import Workbook
from loguru import logger
from openpyxl.styles import PatternFill
from dataclasses import dataclass
from collections import defaultdict

from openpyxl.worksheet.worksheet import Worksheet

from utils.utils import save_dataframe_to_excel, load_data, create_valid_sheet_name
from ensemble_processor import EnsembleProcessor
from utils.visualization_utils import plot_rewards_trend
from reward_func import RewardCalculator, HumanScoreInfo, RewardMetrics
from conf.config import OUTPUT_DIR, SAMPLE_SIZES

# 常量定义
REWARD_PREFIX = "reward"
ENSEMBLE_PREFIX = "ensemble"
SAMPLE_PREFIX = "sample"


@dataclass
class SampleInfo:
    """样本信息数据类"""
    result_df: pd.DataFrame
    human_info: HumanScoreInfo
    reward_metrics: Dict[str, RewardMetrics]
    ensemble_info: Dict[str, Any]
    
    def to_dict(self) -> Dict[str, Any]:
        """转换为字典格式"""
        reward_dict = {}
        for num, metrics in self.reward_metrics.items():
            if isinstance(metrics, RewardMetrics):  # 确保是 RewardMetrics 实例
                reward_dict.update(metrics.to_dict(f"{REWARD_PREFIX}{num}"))
        
        return {
            **self.human_info.to_dict(),
            **reward_dict,
            **self.ensemble_info
        }


class RewardAnalyzer:
    """奖励分析器类，负责分析奖励函数的效果"""
    
    def __init__(self, df: pd.DataFrame):
        """
        初始化奖励分析器
        
        参数:
            df: 包含所有数据的DataFrame
        """
        self.total_requirements = 0
        self.info_result_dict = []
        self.error_wb = Workbook()
        self.df = df
        self.ensemble_processor = EnsembleProcessor()
        
        # 缓存常用属性
        self.reward_numbers = self._get_reward_numbers()
        self.ensemble_headers = self.ensemble_processor.get_ensemble_headers()
        self.headers = self._get_headers()
        
        # 按traj_cnt分组的统计数据
        self.traj_cnt_stats = {}
        
        # 移除默认工作表
        if "Sheet" in self.error_wb.sheetnames:
            self.error_wb.remove(self.error_wb["Sheet"])
    
    def _get_headers(self) -> List[str]:
        """获取表头列表"""
        return [
            "Case_Name",
            *[f"Reward{num}_Score" for num in self.reward_numbers],
            *self.ensemble_headers,
            "Avg_Human_Score",
            "Best_Human_Score",
            "Human_Scores"
        ]
    
    def _collect_row_data(self, size: int) -> Tuple[List[List[Any]], Dict[str, List[float]]]:
        """收集所有行数据和数值统计"""
        sample_key = f'{SAMPLE_PREFIX}_{size}'
        numerical_data = {h: [] for h in self.headers if h not in ["Case_Name", "Human_Scores"]}
        row_data = []
        
        for info in self.info_result_dict:
            if sample_key not in info:
                continue
            
            sample_info = info[sample_key]
            
            # 收集基础数据
            row_values = {
                "Case_Name": info['case_name'],
                **{f"Reward{num}_Score": sample_info[f"select_score_reward{num}"]
                   for num in self.reward_numbers},
                **{header: score for header, score in zip(
                    self.ensemble_headers,
                    self.ensemble_processor.get_ensemble_scores(sample_info)
                )},
                "Avg_Human_Score": sample_info['avg_human_score'],
                "Best_Human_Score": sample_info['best_human_score'],
                "Human_Scores": str(sample_info['human_scores'])
            }
            
            # 按headers顺序构建行数据
            row = [row_values.get(header, '') for header in self.headers]
            row_data.append(row)
            
            # 收集数值数据
            for header, value in row_values.items():
                if header in numerical_data:
                    numerical_data[header].append(float(value))
        
        return row_data, numerical_data
    
    def _create_sample_sheet(self, wb: Workbook, size: int) -> None:
        """创建单个样本大小的工作表"""
        ws = wb.create_sheet(f"S{size}")
        ws.append(self.headers)
        
        row_data, numerical_data = self._collect_row_data(size)
        
        # 写入数据并高亮处理
        highlight_fill = PatternFill(start_color='FEFF00', end_color='FEFF00', fill_type='solid')
        reward4_col_idx = self.headers.index("Reward4_Score") + 1
        avg_human_col_idx = self.headers.index("Avg_Human_Score") + 1
        
        for idx, row in enumerate(row_data, start=2):
            ws.append(row)
            self._highlight_row_if_needed(ws, idx, row, reward4_col_idx, avg_human_col_idx, highlight_fill)
        
        ws.append([])  # 添加空行
        self._append_average_row(ws, numerical_data)
    
    def _append_average_row(self, ws: Worksheet, numerical_data: Dict[str, List[float]]) -> None:
        """添加均值行"""
        mean_row = ["Average"]
        mean_row.extend(
            f"{np.mean(numerical_data[h]):.4f}" if h in numerical_data and numerical_data[h] else ""
            for h in self.headers[1:]
        )
        ws.append(mean_row)
    
    def _calculate_reward_metrics(
            self,
            sampled_group: pd.DataFrame,
            human_info: HumanScoreInfo,
            method_name: str
    ) -> RewardMetrics:
        """计算单个奖励方法的指标"""
        method = getattr(RewardCalculator, method_name)
        reward = method(sampled_group)
        best_idx = reward.idxmax()
        
        return RewardMetrics(
            reward=reward,
            best_idx=best_idx,
            select_score=sampled_group.loc[best_idx, 'human_score'],
            best1_correct=sampled_group.loc[best_idx, 'human_score'] == human_info.best_score,
            best3_correct=sampled_group.loc[best_idx, 'human_score'] in human_info.top3_scores,
            human_rank=human_info.ranks[best_idx]
        )
    
    def _update_result_df(
            self,
            df: pd.DataFrame,
            reward_num: str,
            metrics: RewardMetrics
    ) -> pd.DataFrame:
        """更新结果DataFrame"""
        prefix = f"{REWARD_PREFIX}{reward_num}"
        df[f"{prefix}"] = metrics.reward
        df[f"{prefix}_score"] = metrics.select_score
        df[f"{prefix}_rank"] = metrics.reward.rank(ascending=False, method='min')
        df[f"is_{prefix}_best"] = False
        df.loc[metrics.best_idx, f"is_{prefix}_best"] = True
        return df
    
    def _get_reward_info(self, group: pd.DataFrame, sample_size: int) -> SampleInfo:
        """获取单个样本的奖励分析信息"""
        sampled_group = group.sample(n=sample_size, random_state=49)
        result_df = sampled_group.copy()
        
        # 获取人类评分信息
        human_info = HumanScoreInfo.from_scores(
            pd.to_numeric(sampled_group['human_score'], errors='coerce')
        )
        
        # 计算各种奖励值
        reward_metrics = {
            method_name.replace('calculate_reward', ''): self._calculate_reward_metrics(
                sampled_group, human_info, method_name
            )
            for method_name in self._get_reward_methods()
        }
        
        # 更新结果DataFrame
        for num, metrics in reward_metrics.items():
            result_df = self._update_result_df(result_df, num, metrics)
        
        # 添加集成奖励结果
        result_df = self.ensemble_processor.process_ensemble_rewards(
            result_df, sampled_group, reward_metrics
        )
        ensemble_info = self.ensemble_processor.analyze_ensemble_rewards(
            result_df, human_info.best_score, human_info.top3_scores
        )
        
        return SampleInfo(
            result_df=result_df,
            human_info=human_info,
            reward_metrics=reward_metrics,
            ensemble_info=ensemble_info
        )
    
    def analyze_group(self, group: pd.DataFrame) -> Dict[str, Any]:
        """分析单个组的数据"""
        self.total_requirements += 1
        info = {'case_name': group['case_name'].iloc[0]}
        
        if len(group) < 2:
            return info
        
        for sample_size in SAMPLE_SIZES:
            if len(group) >= sample_size:
                sample_info = self._get_reward_info(group, sample_size)
                info[f"{SAMPLE_PREFIX}_{sample_size}"] = sample_info.to_dict()
                
                sheet_name = create_valid_sheet_name(
                    f"{info['case_name']}_s{sample_size}",
                    self.error_wb.sheetnames
                )
                ws = self.error_wb.create_sheet(sheet_name)
                save_dataframe_to_excel(sample_info.result_df, ws)
        
        return info
    
    def save_case_results(self, filename: str = OUTPUT_DIR / 'case_results.xlsx') -> None:
        """
        保存每个case的详细结果，并计算每列的均值
        高亮显示 reward4 分数小于 human_score 的行
        """
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        for size in SAMPLE_SIZES:
            self._create_sample_sheet(wb, size)
        
        wb.save(filename)
        logger.success(f"Case结果已保存到 '{filename}'")
    
    def _initialize_sample_stats(self, reward_numbers: List[str]) -> Dict[str, Dict]:
        """
        初始化样本统计数据结构
        
        参数:
            reward_numbers: 所有奖励函数编号列表（包括基础奖励和集成奖励）
            
        返回:
            初始化的统计数据字典
        """
        sample_stats = {}
        for size in SAMPLE_SIZES:
            sample_key = f'sample_{size}'
            
            # 区分基础奖励和集成奖励
            base_rewards = [num for num in reward_numbers if not num.startswith('ensemble')]
            ensemble_rewards = [num for num in reward_numbers if num.startswith('ensemble')]
            
            stats_dict = {
                'human_scores': [],
                'best1_correct_counts': {},  # 将在下面填充
                'best3_correct_counts': {},  # 将在下面填充
                'reward_select_human_score': {},  # 将在下面填充
                'human_ranks': {},  # 将在下面填充
                'total_samples': 0
            }
            
            # 初始化基础奖励的统计项
            for num in base_rewards:
                reward_key = f'reward{num}'
                stats_dict['best1_correct_counts'][reward_key] = 0
                stats_dict['best3_correct_counts'][reward_key] = 0
                stats_dict['reward_select_human_score'][f'reward_select_human_score_{reward_key}'] = []
                stats_dict['human_ranks'][reward_key] = []
            
            # 初始化集成奖励的统计项
            for num in ensemble_rewards:
                stats_dict['best1_correct_counts'][num] = 0
                stats_dict['best3_correct_counts'][num] = 0
                stats_dict['reward_select_human_score'][f'reward_select_human_score_{num}'] = []
                stats_dict['human_ranks'][num] = []
            
            sample_stats[sample_key] = stats_dict
        
        return sample_stats
    
    def _get_reward_methods(self) -> List[str]:
        """
        获取所有奖励计算方法名称
        
        返回:
            方法名称列表
        """
        return [method for method in dir(RewardCalculator)
                if method.startswith('calculate_reward')]
    
    def _get_reward_numbers(self) -> List[str]:
        """
        获取所有奖励函数编号
        
        返回:
            编号列表
        """
        return [method.replace('calculate_reward', '')
                for method in self._get_reward_methods()]
    
    def visualize_rewards_trend(self, save_path: str = OUTPUT_DIR / 'reward_trends.png'):
        """使用折线图展示不同sample size下各个reward方法的表现趋势"""
        trend_data = defaultdict(list)
        human_scores = {
            'avg': {size: [] for size in SAMPLE_SIZES},
            'best': {size: [] for size in SAMPLE_SIZES}
        }
        
        for size in SAMPLE_SIZES:
            sample_key = f'{SAMPLE_PREFIX}_{size}'
            reward_scores = defaultdict(list)
            
            for info in self.info_result_dict:
                if sample_key not in info:
                    continue
                
                sample_info = info[sample_key]
                
                # 收集reward和ensemble分数
                
                for header in self.headers:
                    if header in ['Case_Name', 'Human_Scores', 'Avg_Human_Score', 'Best_Human_Score']:
                        continue
                    mapped_name = header.split("_")[0].lower()
                    reward_scores[header].append(sample_info.get(f"select_score_{mapped_name}", 0))
                
                # 收集human分数
                human_scores['avg'][size].append(sample_info['avg_human_score'])
                human_scores['best'][size].append(sample_info['best_human_score'])
            
            # 计算统计值
            for method, scores in reward_scores.items():
                if scores:
                    trend_data['sample_size'].append(size)
                    trend_data['reward_method'].append(method)
                    trend_data['mean_score'].append(np.mean(scores))
                    trend_data['std_score'].append(np.std(scores))
        
        # 调用可视化工具绘制图表
        plot_rewards_trend(
            trend_data=dict(trend_data),
            reward_numbers=self.reward_numbers,
            sample_sizes=SAMPLE_SIZES,
            avg_human_scores={k: np.mean(v) if v else 0 for k, v in human_scores['avg'].items()},
            best_human_scores={k: np.mean(v) if v else 0 for k, v in human_scores['best'].items()},
            save_path=save_path
        )
    
    def _highlight_row_if_needed(
            self,
            ws: Worksheet,
            row_idx: int,
            row_data: List[Any],
            reward4_col_idx: int,
            avg_human_col_idx: int,
            highlight_fill: PatternFill
    ) -> None:
        """如果需要则高亮整行"""
        try:
            reward4_value = float(row_data[reward4_col_idx - 1])
            avg_human_value = float(row_data[avg_human_col_idx - 1])
            
            if reward4_value < avg_human_value:
                for cell in ws[row_idx]:
                    cell.fill = highlight_fill
        except (ValueError, IndexError):
            logger.warning(
                f"无法处理行 {row_idx} 的高亮: reward4={row_data[reward4_col_idx - 1]}, avg_human={row_data[avg_human_col_idx - 1]}")


def main():
    """主函数"""
    # 配置日志
    logger.add("reward_analysis.log", rotation="10 MB", level="INFO")
    logger.info("开始分析奖励函数...")
    
    # 加载数据
    df = load_data()
    logger.info(f"已加载数据，共 {len(df)} 条记录")
    
    # 按case_name分组
    grouped = df.groupby('case_name')
    logger.info(f"数据已分组，共 {len(grouped)} 个需求组")
    
    # 初始化分析器
    analyzer = RewardAnalyzer(df)
    
    # 处理每个需求组
    for requirement, group in grouped:
        info = analyzer.analyze_group(group)
        analyzer.info_result_dict.append(info)
    
    analyzer.error_wb.save(OUTPUT_DIR / 'best1_errors.xlsx')
    
    # 保存详细结果
    analyzer.save_case_results()
    
    # 分析并可视化reward结果
    analyzer.visualize_rewards_trend()
    
    logger.success("分析完成！")


if __name__ == "__main__":
    main()
