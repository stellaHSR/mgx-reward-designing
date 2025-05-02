import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

version_df = pd.read_excel("version.xlsx")
versions = version_df["执行版本"].values.tolist()
# Load the Excel file
df = pd.read_excel('mgx_ensembling.xlsx')
df = df[df["执行版本"].apply(lambda x: x in versions)]
# Group by requirement to process each task separately
# grouped = df.groupby('requirement')
grouped = df.groupby('case_name')

# Initialize counters for accuracy statistics
total_requirements = 0
reward1_best1_correct = 0
reward1_best5_correct = 0
reward2_best1_correct = 0
reward2_best5_correct = 0

# Create a new workbook for saving incorrect predictions
error_wb = Workbook()
# Remove the default sheet
if "Sheet" in error_wb.sheetnames:
    error_wb.remove(error_wb["Sheet"])

info_result_dict = []
# Process each requirement group
for requirement, group in grouped:
    total_requirements += 1
    info = {"requirement": requirement,
            "scenario":group["scenario"].values[0],
            "traj_cnt":len(group),
            "human_score": -1.0,
            "select_score_1": -1.0,
            "select_score_2": -1.0,
            }
    
    # Skip if group has less than 2 entries (need multiple candidates to compare)
    if len(group) < 2:
        continue
    
    # Calculate Reward 1
    reward1 = group['executability'] * (
            (0.1102 * group['code_lines']) +
            (0.0488 * group['code_files_count']) +
            # (0.1525 * group['completion_cost']) +
            # (0.1559 * group['cost($)']) +
            # (0.2086 * group['completion_tokens']) +
            (0.2093 * group['interactions_count']) +
            (0.0492 * group['llm_score'])
        # (0.0656 * group['total_time'])
    )
    
    # Calculate Reward 2
    max_cost = group['cost($)'].max()
    max_code_lines = group['code_lines'].max()
    
    reward2 = group['executability'] * (
            group['total_time'] * 1
        #     # (group['cost($)'] / max_cost if max_cost > 0 else 0) * 0.5 +
        #     # (group['code_lines'] / max_code_lines if max_code_lines > 0 else 0) * 0.2
    )
    
    # reward2 = group['code_lines'] *1
    #     # (group['cost($)'] / max_cost if max_cost > 0 else 0) * 0.5 +
    #     # (group['code_lines'] / max_code_lines if max_code_lines > 0 else 0) * 0.2
    # # )
    
    reward2 = group['executability'] * (
            group['llm_score'] +
            (group['cost($)'] / max_cost if max_cost > 0 else 0) * 0.5 +
            (group['code_lines'] / max_code_lines if max_code_lines > 0 else 0) * 0.2
    )
    
    # reward2 = group['llm_score'] * group['cost($)'] * group['total_time']
    
    # Find best candidates according to rewards
    reward1_best_idx = reward1.idxmax()
    reward2_best_idx = reward2.idxmax()
    
    # Get human scores and sort to find top human-rated candidates
    human_scores = group['human_score'].sort_values(ascending=False)
    select_reward_score1 = group.loc[reward1_best_idx, 'human_score']
    select_reward_score2 = group.loc[reward2_best_idx, 'human_score']
    
    best_human_score = human_scores.iloc[0]
    top5_human_scores = set(human_scores.iloc[:5].values)

    info["human_score"] = np.mean(human_scores)
    info["select_score_1"] = select_reward_score1
    info["select_score_2"] = select_reward_score2
    
    print(info)
    
    info_result_dict.append(info)
    # Check if reward predictions match human evaluations
    reward1_best1_is_correct = group.loc[reward1_best_idx, 'human_score'] == best_human_score
    reward2_best1_is_correct = group.loc[reward2_best_idx, 'human_score'] == best_human_score
    
    if reward1_best1_is_correct:
        reward1_best1_correct += 1
    
    if group.loc[reward1_best_idx, 'human_score'] in top5_human_scores:
        reward1_best5_correct += 1
    
    if reward2_best1_is_correct:
        reward2_best1_correct += 1
    
    if group.loc[reward2_best_idx, 'human_score'] in top5_human_scores:
        reward2_best5_correct += 1
    
    # Save the data if either reward formula failed to identify the best entry
    if not (reward1_best1_is_correct and reward2_best1_is_correct):
        # Create a copy of the group dataframe with additional reward columns
        result_df = group.copy()
        result_df = result_df[['case_name_uni', 'product', 'scenario', 'llm_score',
                               'human_score', 'executability', 'quality', 'interactions_count', 'cost($)', 'case_name']]
        result_df['reward1'] = reward1
        result_df['reward2'] = reward2
        result_df['reward1_score'] = select_reward_score1
        result_df['reward2_score'] = select_reward_score2
        
        # Add columns to identify the selected candidates
        result_df['is_reward1_best'] = False
        result_df['is_reward2_best'] = False
        result_df.loc[reward1_best_idx, 'is_reward1_best'] = True
        result_df.loc[reward2_best_idx, 'is_reward2_best'] = True
        
        # Add a column to identify entries with the highest human score
        result_df['is_human_best'] = result_df['human_score'] == best_human_score
        
        # Sort by human_score descending to put the best human-rated entries at the top
        result_df = result_df.sort_values(by='human_score', ascending=False)
        
        # result_df = result_df[[""]]
        # Create a sheet name based on the requirement (truncate if too long)
        sheet_name = str(requirement)[:30]  # Excel sheet names have a 31 character limit
        # Replace invalid characters for Excel sheet names
        sheet_name = sheet_name.replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_').replace('[',
                                                                                                                 '_').replace(
            ']', '_').replace(':', '_')
        
        # Create a sheet with a unique name if there are duplicates
        i = 1
        base_sheet_name = sheet_name
        while sheet_name in error_wb.sheetnames:
            sheet_name = f"{base_sheet_name}_{i}"
            i += 1
        
        # Create new sheet and add data
        ws = error_wb.create_sheet(sheet_name)
        
        # Write the DataFrame to the worksheet
        for r in dataframe_to_rows(result_df, index=False, header=True):
            ws.append(r)
    
# Save the workbook with incorrect predictions
error_wb.save('best1_errors.xlsx')

# Calculate accuracy statistics
if total_requirements > 0:
    avg_select_score_1 = np.mean([i["select_score_1"] for i in info_result_dict])
    avg_select_score_2 =  np.mean([i["select_score_2"] for i in info_result_dict])
    avg_human_score =  np.mean([i["human_score"] for i in info_result_dict])
    print(f"\n Avg reward1_best1: {avg_select_score_1}")
    print(f"\n Avg reward2_best1: {avg_select_score_2}")
    print(f"\n Avg human_score: {avg_human_score}")
    
    reward1_best1_accuracy = reward1_best1_correct / total_requirements
    reward1_best5_accuracy = reward1_best5_correct / total_requirements
    reward2_best1_accuracy = reward2_best1_correct / total_requirements
    reward2_best5_accuracy = reward2_best5_correct / total_requirements
    
    
    print(f"Total requirements analyzed: {total_requirements}")
    print("\nReward 1 (Formula from task description):")
    print(f"Best@1 Accuracy: {reward1_best1_accuracy:.4f} ({reward1_best1_correct}/{total_requirements})")
    print(f"Best@5 Accuracy: {reward1_best5_accuracy:.4f} ({reward1_best5_correct}/{total_requirements})")
    
    print("\nReward 2 (Custom formula):")
    print(f"Best@1 Accuracy: {reward2_best1_accuracy:.4f} ({reward2_best1_correct}/{total_requirements})")
    print(f"Best@5 Accuracy: {reward2_best5_accuracy:.4f} ({reward2_best5_correct}/{total_requirements})")
    
    print(f"\nIncorrect Best@1 predictions have been saved to 'best1_errors.xlsx'")
else:
    print("No requirement groups found for analysis.")
